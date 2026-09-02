# import os
# import csv
# from datetime import datetime
# from openpyxl import Workbook, load_workbook
# from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
# from openpyxl.utils import get_column_letter


# class EnduranceDataLogger:

#     # ==========================================================
#     # PARAMETERS TO LOG
#     # ==========================================================

#     COMMON_PARAMETERS = [

#         # OBC
#         "OBC Input Voltage",
#         "OBC Input Current",
#         "OBC Output Voltage",
#         "OBC Output Current",
#         "OBC_Input_Power",
#         "OBC_Output_Power",
#         "OBC Efficiency",

#         # HPDCDC
#         "HPDCDC Input Voltage",
#         "HPDCDC Input Current",
#         "HPDCDC Output Voltage",
#         "HPDCDC Output Current",
#         "HPDCDC_Input_Power",
#         "HPDCDC_Output_Power",
#         "HPDCDC_Efficiency",

#         # Temperature
#         "OBC_TEMP",
#         "OBC_FET_TEMP",
#         "HPDCDC_TEMP",
#     ]

#     # ==========================================================
#     # INITIALIZATION
#     # ==========================================================

#     def __init__(self, base_folder="CSV_Logs", report_folder="Report"):

#         self.base_folder = base_folder
#         self.report_folder = report_folder

#         # Current test run
#         self.test_folder_name = None
#         self.test_name = None

#         # CSV
#         self.csv_files = {}
#         self.csv_writers = {}

#         # Excel
#         self.excel_files = {}
#         self.excel_rows = {}

#         # Start time for each DUT/cycle
#         self.cycle_start_time = {}

#     # ==========================================================
#     # SANITIZE FILE/FOLDER NAME
#     # ==========================================================

#     def sanitize_name(self, name):

#         if name is None:
#             name = "Unknown_Test"

#         invalid = '<>:"/\\|?*'

#         for char in invalid:
#             name = name.replace(char, "_")

#         return name.strip()

#     # ==========================================================
#     # START TEST
#     # ==========================================================

#     def start_test(self, test_name):

#         self.test_name = self.sanitize_name(test_name)

#         timestamp = datetime.now().strftime(
#             "%Y-%m-%d_%H-%M-%S"
#         )

#         self.test_folder_name = (
#             f"{self.test_name}_{timestamp}"
#         )

#         # ------------------------------------------------------
#         # CSV ROOT
#         # ------------------------------------------------------

#         csv_root = os.path.join(
#             self.base_folder,
#             self.test_folder_name
#         )

#         os.makedirs(
#             csv_root,
#             exist_ok=True
#         )

#         # ------------------------------------------------------
#         # EXCEL ROOT
#         # ------------------------------------------------------

#         excel_root = os.path.join(
#             self.report_folder,
#             "Endurance",
#             self.test_folder_name
#         )

#         os.makedirs(
#             excel_root,
#             exist_ok=True
#         )

#         print(
#             f"Data logging started:\n"
#             f"CSV    : {csv_root}\n"
#             f"Report : {excel_root}"
#         )

#     # ==========================================================
#     # START CYCLE
#     # ==========================================================

#     def start_cycle(
#         self,
#         test_name,
#         dut_no,
#         cycle_no,
#         serial_number="",
#         user_name="",
#         temperature_type="",
#         initial_temperature="",
#         source_type=""
#     ):

#         # If start_test() was not called
#         if self.test_folder_name is None:

#             self.start_test(test_name)

#         # ------------------------------------------------------
#         # DUT FOLDER - CSV
#         # ------------------------------------------------------

#         dut_csv_folder = os.path.join(
#             self.base_folder,
#             self.test_folder_name,
#             f"DUT{dut_no}"
#         )

#         os.makedirs(
#             dut_csv_folder,
#             exist_ok=True
#         )

#         csv_path = os.path.join(
#             dut_csv_folder,
#             f"Cycle_{cycle_no}.csv"
#         )

#         # ------------------------------------------------------
#         # DUT FOLDER - EXCEL
#         # ------------------------------------------------------

#         dut_excel_folder = os.path.join(
#             self.report_folder,
#             "Endurance",
#             self.test_folder_name,
#             f"DUT{dut_no}"
#         )

#         os.makedirs(
#             dut_excel_folder,
#             exist_ok=True
#         )

#         excel_path = os.path.join(
#             dut_excel_folder,
#             f"Cycle_{cycle_no}.xlsx"
#         )

#         # ------------------------------------------------------
#         # START TIME
#         # ------------------------------------------------------

#         start_time = datetime.now()

#         self.cycle_start_time[
#             (dut_no, cycle_no)
#         ] = start_time

#         # ======================================================
#         # CSV
#         # ======================================================

#         csv_file = open(
#             csv_path,
#             "w",
#             newline="",
#             encoding="utf-8"
#         )

#         headers = [
#             "Timestamp",
#             "Time_sec",
#             "Mode"
#         ]

#         for parameter in self.COMMON_PARAMETERS:

#             headers.append(
#                 f"CAN_{parameter}"
#             )

#         for parameter in self.COMMON_PARAMETERS:

#             headers.append(
#                 f"HW_{parameter}"
#             )

#         writer = csv.DictWriter(
#             csv_file,
#             fieldnames=headers
#         )

#         writer.writeheader()

#         key = (
#             dut_no,
#             cycle_no
#         )

#         self.csv_files[key] = csv_file
#         self.csv_writers[key] = writer

#         # ======================================================
#         # EXCEL
#         # ======================================================

#         workbook = Workbook()

#         worksheet = workbook.active

#         worksheet.title = f"Cycle {cycle_no}"

#         # ------------------------------------------------------
#         # Store Excel information
#         # ------------------------------------------------------

#         self.excel_files[key] = {
#             "workbook": workbook,
#             "worksheet": worksheet,
#             "path": excel_path,
#             "charging_rows": [],
#             "discharging_rows": [],
#             "charging_start_row": None,
#             "discharging_start_row": None,
#             "charging_header_row": None,
#             "discharging_header_row": None,
#             "charging_average_row": None,
#             "discharging_average_row": None,
#         }

#         self._create_excel_header(
#             key=key,
#             serial_number=serial_number,
#             dut_no=dut_no,
#             start_time=start_time,
#             user_name=user_name,
#             temperature_type=temperature_type,
#             initial_temperature=initial_temperature,
#             source_type=source_type
#         )

#         workbook.save(excel_path)

#         print(
#             f"CSV started   : {csv_path}"
#         )

#         print(
#             f"Excel started : {excel_path}"
#         )

#     # ==========================================================
#     # EXCEL HEADER
#     # ==========================================================

#     def _create_excel_header(
#         self,
#         key,
#         serial_number,
#         dut_no,
#         start_time,
#         user_name,
#         temperature_type,
#         initial_temperature,
#         source_type
#     ):

#         info = self.excel_files[key]

#         ws = info["worksheet"]

#         # ------------------------------------------------------
#         # TITLE
#         # ------------------------------------------------------

#         ws["A1"] = "Endurance Test Report"

#         ws["A1"].font = Font(
#             bold=True,
#             size=16
#         )

#         ws.merge_cells(
#             start_row=1,
#             start_column=1,
#             end_row=1,
#             end_column=6
#         )

#         # ------------------------------------------------------
#         # INFORMATION
#         # ------------------------------------------------------

#         headers = [
#             "Serial Number",
#             "DUT Slot",
#             "Start Time",
#             "User Name",
#             "Temperature Type",
#             "Initial Temperature(°C)",
#             "Source Type"
#         ]

#         values = [
#             serial_number,
#             f"DUT{dut_no}",
#             start_time.strftime(
#                 "%Y-%m-%d %H:%M:%S"
#             ),
#             user_name,
#             temperature_type,
#             initial_temperature,
#             source_type
#         ]

#         # Header row
#         for column, value in enumerate(
#             headers,
#             start=1
#         ):

#             cell = ws.cell(
#                 row=3,
#                 column=column
#             )

#             cell.value = value
#             cell.font = Font(
#                 bold=True
#             )

#         # Values row
#         for column, value in enumerate(
#             values,
#             start=1
#         ):

#             ws.cell(
#                 row=4,
#                 column=column
#             ).value = value

#         # ------------------------------------------------------
#         # Formatting
#         # ------------------------------------------------------

#         for row in ws.iter_rows(
#             min_row=3,
#             max_row=4,
#             min_col=1,
#             max_col=len(headers)
#         ):

#             for cell in row:

#                 cell.alignment = Alignment(
#                     horizontal="center",
#                     vertical="center"
#                 )

#         # ------------------------------------------------------
#         # Start Charging section
#         # ------------------------------------------------------

#         row = 6

#         ws.cell(
#             row=row,
#             column=1
#         ).value = "OBC Charging"

#         ws.cell(
#             row=row,
#             column=1
#         ).font = Font(
#             bold=True,
#             size=14
#         )

#         info["charging_start_row"] = row

#         row += 1

#         headers = self._get_data_headers()

#         info["charging_header_row"] = row

#         self._write_data_headers(
#             ws,
#             row,
#             headers
#         )

#     # ==========================================================
#     # DATA HEADERS
#     # ==========================================================

#     def _get_data_headers(self):

#         headers = [
#             "Timestamp",
#             "Time_sec",
#             "Mode"
#         ]

#         for parameter in self.COMMON_PARAMETERS:

#             headers.append(
#                 f"CAN_{parameter}"
#             )

#         for parameter in self.COMMON_PARAMETERS:

#             headers.append(
#                 f"HW_{parameter}"
#             )

#         return headers

#     # ==========================================================
#     # WRITE DATA HEADERS
#     # ==========================================================

#     def _write_data_headers(
#         self,
#         worksheet,
#         row,
#         headers
#     ):

#         for column, header in enumerate(
#             headers,
#             start=1
#         ):

#             cell = worksheet.cell(
#                 row=row,
#                 column=column
#             )

#             cell.value = header

#             cell.font = Font(
#                 bold=True
#             )

#             cell.alignment = Alignment(
#                 horizontal="center"
#             )

#     # ==========================================================
#     # LOG DATA
#     # ==========================================================

#     def log_data(
#         self,
#         dut_no,
#         cycle_no,
#         mode,
#         can_values,
#         hardware_values,
#         timestamp=None,
#         time_sec=None
#     ):

#         key = (
#             dut_no,
#             cycle_no
#         )

#         # ------------------------------------------------------
#         # Validate
#         # ------------------------------------------------------

#         if key not in self.csv_writers:

#             print(
#                 f"Logger not started for "
#                 f"DUT{dut_no} Cycle {cycle_no}"
#             )

#             return

#         # ------------------------------------------------------
#         # Timestamp
#         # ------------------------------------------------------

#         if timestamp is None:

#             timestamp = datetime.now()

#         if isinstance(timestamp, datetime):

#             timestamp_text = timestamp.strftime(
#                 "%Y-%m-%d %H:%M:%S"
#             )

#         else:

#             timestamp_text = str(timestamp)

#         # ------------------------------------------------------
#         # Time seconds
#         # ------------------------------------------------------

#         if time_sec is None:

#             start = self.cycle_start_time.get(
#                 key,
#                 timestamp
#             )

#             if isinstance(start, datetime) and isinstance(timestamp, datetime):

#                 time_sec = (
#                     timestamp - start
#                 ).total_seconds()

#             else:

#                 time_sec = 0

#         # ======================================================
#         # BUILD ROW
#         # ======================================================

#         row = {
#             "Timestamp": timestamp_text,
#             "Time_sec": round(
#                 float(time_sec),
#                 2
#             ),
#             "Mode": mode
#         }

#         # ------------------------------------------------------
#         # CAN
#         # ------------------------------------------------------

#         for parameter in self.COMMON_PARAMETERS:

#             value = can_values.get(
#                 parameter,
#                 ""
#             )

#             row[
#                 f"CAN_{parameter}"
#             ] = value

#         # ------------------------------------------------------
#         # HARDWARE
#         # ------------------------------------------------------

#         for parameter in self.COMMON_PARAMETERS:

#             value = hardware_values.get(
#                 parameter,
#                 ""
#             )

#             row[
#                 f"HW_{parameter}"
#             ] = value

#         # ======================================================
#         # CSV
#         # ======================================================

#         writer = self.csv_writers[key]

#         writer.writerow(row)

#         self.csv_files[key].flush()

#         # ======================================================
#         # EXCEL
#         # ======================================================

#         info = self.excel_files[key]

#         ws = info["worksheet"]

#         # ------------------------------------------------------
#         # Determine section
#         # ------------------------------------------------------

#         mode_lower = str(
#             mode
#         ).lower()

#         if mode_lower in (
#             "charge",
#             "charging"
#         ):

#             section = "charging"

#         elif mode_lower in (
#             "discharge",
#             "discharging"
#         ):

#             section = "discharging"

#         else:

#             # REST is not stored in either section
#             return

#         # ------------------------------------------------------
#         # Create DCDC section when first needed
#         # ------------------------------------------------------

#         if (
#             section == "discharging"
#             and info["discharging_header_row"] is None
#         ):

#             self._create_discharging_section(
#                 key
#             )

#         # ------------------------------------------------------
#         # Determine row
#         # ------------------------------------------------------

#         if section == "charging":

#             header_row = info[
#                 "charging_header_row"
#             ]

#             data_rows = info[
#                 "charging_rows"
#             ]

#         else:

#             header_row = info[
#                 "discharging_header_row"
#             ]

#             data_rows = info[
#                 "discharging_rows"
#             ]

#         row_number = (
#             header_row +
#             1 +
#             len(data_rows)
#         )

#         # ------------------------------------------------------
#         # Write Excel row
#         # ------------------------------------------------------

#         values = []

#         for header in self._get_data_headers():

#             values.append(
#                 row.get(
#                     header,
#                     ""
#                 )
#             )

#         for column, value in enumerate(
#             values,
#             start=1
#         ):

#             ws.cell(
#                 row=row_number,
#                 column=column
#             ).value = value

#         data_rows.append(
#             row_number
#         )

#         # ------------------------------------------------------
#         # Auto width
#         # ------------------------------------------------------

#         self._adjust_column_widths(
#             ws
#         )

#         # Save
#         info["workbook"].save(
#             info["path"]
#         )

#     # ==========================================================
#     # CREATE DISCHARGING SECTION
#     # ==========================================================

#     def _create_discharging_section(
#         self,
#         key
#     ):

#         info = self.excel_files[key]

#         ws = info["worksheet"]

#         # ------------------------------------------------------
#         # Find next available row
#         # ------------------------------------------------------

#         max_row = ws.max_row

#         row = max_row + 3

#         # ------------------------------------------------------
#         # Heading
#         # ------------------------------------------------------

#         ws.cell(
#             row=row,
#             column=1
#         ).value = "DCDC Discharging"

#         ws.cell(
#             row=row,
#             column=1
#         ).font = Font(
#             bold=True,
#             size=14
#         )

#         info["discharging_start_row"] = row

#         row += 1

#         # ------------------------------------------------------
#         # Column headers
#         # ------------------------------------------------------

#         info["discharging_header_row"] = row

#         self._write_data_headers(
#             ws,
#             row,
#             self._get_data_headers()
#         )

#     # ==========================================================
#     # FINISH CYCLE
#     # ==========================================================

#     def finish_cycle(
#         self,
#         dut_no,
#         cycle_no
#     ):

#         key = (
#             dut_no,
#             cycle_no
#         )

#         # ======================================================
#         # CSV
#         # ======================================================

#         if key in self.csv_files:

#             try:

#                 self.csv_files[key].flush()
#                 self.csv_files[key].close()

#             except Exception as e:

#                 print(
#                     f"CSV close error: {e}"
#                 )

#             del self.csv_files[key]
#             del self.csv_writers[key]

#             print(
#                 f"CSV closed: "
#                 f"DUT{dut_no}, "
#                 f"Cycle {cycle_no}"
#             )

#         # ======================================================
#         # EXCEL
#         # ======================================================

#         if key in self.excel_files:

#             info = self.excel_files[key]

#             ws = info["worksheet"]

#             # --------------------------------------------------
#             # Charging average
#             # --------------------------------------------------

#             if info["charging_rows"]:

#                 self._write_average_row(
#                     ws,
#                     info["charging_rows"],
#                     "charging"
#                 )

#             # --------------------------------------------------
#             # Discharging average
#             # --------------------------------------------------

#             if info["discharging_rows"]:

#                 self._write_average_row(
#                     ws,
#                     info["discharging_rows"],
#                     "discharging"
#                 )

#             # --------------------------------------------------
#             # Formatting
#             # --------------------------------------------------

#             self._format_excel(
#                 ws
#             )

#             # --------------------------------------------------
#             # Save
#             # --------------------------------------------------

#             info["workbook"].save(
#                 info["path"]
#             )

#             print(
#                 f"Excel closed: "
#                 f"DUT{dut_no}, "
#                 f"Cycle {cycle_no}"
#             )

#             del self.excel_files[key]

#         # ------------------------------------------------------
#         # Remove start time
#         # ------------------------------------------------------

#         self.cycle_start_time.pop(
#             key,
#             None
#         )

#     # ==========================================================
#     # WRITE AVERAGE
#     # ==========================================================

#     def _write_average_row(
#         self,
#         worksheet,
#         data_rows,
#         section
#     ):

#         if not data_rows:
#             return

#         average_row = (
#             max(data_rows) + 1
#         )

#         worksheet.cell(
#             row=average_row,
#             column=1
#         ).value = "Average"

#         worksheet.cell(
#             row=average_row,
#             column=1
#         ).font = Font(
#             bold=True
#         )

#         # ------------------------------------------------------
#         # Average numeric columns
#         # ------------------------------------------------------

#         headers = self._get_data_headers()

#         for column, header in enumerate(
#             headers,
#             start=1
#         ):

#             # Skip non-numeric
#             if header in (
#                 "Timestamp",
#                 "Mode"
#             ):
#                 continue

#             values = []

#             for row in data_rows:

#                 value = worksheet.cell(
#                     row=row,
#                     column=column
#                 ).value

#                 try:

#                     if value is not None and value != "":

#                         values.append(
#                             float(value)
#                         )

#                 except (
#                     ValueError,
#                     TypeError
#                 ):

#                     pass

#             if values:

#                 average = (
#                     sum(values) /
#                     len(values)
#                 )

#                 worksheet.cell(
#                     row=average_row,
#                     column=column
#                 ).value = round(
#                     average,
#                     4
#                 )

#         # ------------------------------------------------------
#         # Save average row reference
#         # ------------------------------------------------------

#         if section == "charging":

#             # Nothing else required

#             pass

#         elif section == "discharging":

#             pass

#     # ==========================================================
#     # FORMAT EXCEL
#     # ==========================================================

#     def _format_excel(
#         self,
#         worksheet
#     ):

#         thin_border = Border(
#             left=Side(style="thin"),
#             right=Side(style="thin"),
#             top=Side(style="thin"),
#             bottom=Side(style="thin")
#         )

#         # ------------------------------------------------------
#         # Format all used cells
#         # ------------------------------------------------------

#         for row in worksheet.iter_rows():

#             for cell in row:

#                 cell.alignment = Alignment(
#                     vertical="center",
#                     horizontal="center"
#                 )

#         # ------------------------------------------------------
#         # Borders
#         # ------------------------------------------------------

#         for row in worksheet.iter_rows():

#             for cell in row:

#                 if cell.value is not None:

#                     cell.border = thin_border

#         # ------------------------------------------------------
#         # Header rows
#         # ------------------------------------------------------

#         for row in range(
#             1,
#             worksheet.max_row + 1
#         ):

#             value = worksheet.cell(
#                 row=row,
#                 column=1
#             ).value

#             if value in (
#                 "OBC Charging",
#                 "DCDC Discharging"
#             ):

#                 worksheet.cell(
#                     row=row,
#                     column=1
#                 ).font = Font(
#                     bold=True,
#                     size=14
#                 )

#         # ------------------------------------------------------
#         # Freeze panes
#         # ------------------------------------------------------

#         worksheet.freeze_panes = "A8"

#     # ==========================================================
#     # COLUMN WIDTH
#     # ==========================================================

#     def _adjust_column_widths(
#         self,
#         worksheet
#     ):

#         for column_cells in worksheet.columns:

#             max_length = 0

#             column_letter = get_column_letter(
#                 column_cells[0].column
#             )

#             for cell in column_cells:

#                 try:

#                     length = len(
#                         str(cell.value)
#                     )

#                     if length > max_length:

#                         max_length = length

#                 except Exception:

#                     pass

#             worksheet.column_dimensions[
#                 column_letter
#             ].width = min(
#                 max(max_length + 2, 12),
#                 30
#             )

#     # ==========================================================
#     # CLOSE ALL
#     # ==========================================================

#     def close_all(self):

#         # Close CSV files

#         for key, file in list(
#             self.csv_files.items()
#         ):

#             try:
#                 file.flush()
#                 file.close()
#             except Exception:
#                 pass

#         self.csv_files.clear()
#         self.csv_writers.clear()

#         # Save Excel files

#         for key, info in list(
#             self.excel_files.items()
#         ):

#             try:

#                 self._format_excel(
#                     info["worksheet"]
#                 )

#                 info["workbook"].save(
#                     info["path"]
#                 )

#             except Exception as e:

#                 print(
#                     f"Excel close error: {e}"
#                 )

#         self.excel_files.clear()

#         print(
#             "All data loggers closed."
#         )

import os
import csv
import json
import sqlite3
import threading
import queue
import time
import copy

from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter


class EnduranceDataLogger:

    # ==========================================================
    # PARAMETERS TO LOG
    # ==========================================================

    COMMON_PARAMETERS = [

        # ------------------------------------------------------
        # OBC
        # ------------------------------------------------------

        "OBC Input Voltage",
        "OBC Input Current",
        "OBC Output Voltage",
        "OBC Output Current",
        "OBC_Input_Power",
        "OBC_Output_Power",
        "OBC Efficiency",

        # ------------------------------------------------------
        # HPDCDC
        # ------------------------------------------------------

        "HPDCDC Input Voltage",
        "HPDCDC Input Current",
        "HPDCDC Output Voltage",
        "HPDCDC Output Current",
        "HPDCDC_Input_Power",
        "HPDCDC_Output_Power",
        "HPDCDC_Efficiency",

        # ------------------------------------------------------
        # Temperature
        # ------------------------------------------------------

        "OBC_TEMP",
        "OBC_FET_TEMP",
        "HPDCDC_TEMP",
    ]

    PARAMETER_UNITS = {
        "OBC Input Voltage": "V",
        "OBC Input Current": "A",
        "OBC Output Voltage": "V",
        "OBC Output Current": "A",
        "OBC_Input_Power": "W",
        "OBC_Output_Power": "W",
        "OBC Efficiency": "%",
        
        "HPDCDC Input Voltage": "V",
        "HPDCDC Input Current": "A",
        "HPDCDC Output Voltage": "V",
        "HPDCDC Output Current": "A",
        "HPDCDC_Input_Power": "W",
        "HPDCDC_Output_Power": "W",
        "HPDCDC_Efficiency": "%",
        
        "OBC_TEMP": "°C",
        "OBC_FET_TEMP": "°C",
        "HPDCDC_TEMP": "°C"
    }


    # ==========================================================
    # RETRY / SAVE SETTINGS
    # ==========================================================

    RETRY_INTERVAL = 1.0

    EXCEL_SAVE_INTERVAL = 5.0

    # ==========================================================
    # INITIALIZATION
    # ==========================================================

    def __init__(
        self,
        base_folder="CSV_Logs",
        report_folder="Report",
        database_folder="Data"
    ):

        self.base_folder = base_folder

        self.report_folder = report_folder

        self.database_folder = database_folder

        # ------------------------------------------------------
        # Current test
        # ------------------------------------------------------

        self.test_name = None

        self.test_folder_name = None

        # ------------------------------------------------------
        # Cycle start time
        # ------------------------------------------------------

        self.cycle_start_time = {}

        # ------------------------------------------------------
        # CSV information
        # ------------------------------------------------------

        self.csv_paths = {}

        # ------------------------------------------------------
        # Excel information
        # ------------------------------------------------------

        self.excel_files = {}

        # ------------------------------------------------------
        # Queue
        # ------------------------------------------------------

        self.log_queue = queue.Queue()

        # ------------------------------------------------------
        # Thread control
        # ------------------------------------------------------

        self.running = True

        # ------------------------------------------------------
        # SQLite
        # ------------------------------------------------------

        os.makedirs(
            self.database_folder,
            exist_ok=True
        )

        self.db_path = os.path.join(
            self.database_folder,
            "endurance_data.db"
        )

        self._initialize_database()

        # ------------------------------------------------------
        # Start worker
        # ------------------------------------------------------

        self.logger_thread = threading.Thread(
            target=self._logger_worker,
            name="EnduranceDataLogger",
            daemon=True
        )

        self.logger_thread.start()

        print(
            "Endurance logger worker started."
        )

    # ==========================================================
    # SQLITE DATABASE
    # ==========================================================

    def _initialize_database(self):

        conn = sqlite3.connect(
            self.db_path,
            timeout=30
        )

        try:

            cursor = conn.cursor()

            cursor.execute(
                """
                CREATE TABLE IF NOT EXISTS endurance_data
                (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,

                    test_name TEXT,

                    test_folder TEXT,

                    dut_no INTEGER,

                    cycle_no INTEGER,

                    timestamp TEXT,

                    time_sec REAL,

                    mode TEXT,

                    can_values TEXT,

                    hardware_values TEXT,

                    csv_written INTEGER DEFAULT 0,

                    excel_written INTEGER DEFAULT 0,

                    created_at TEXT DEFAULT CURRENT_TIMESTAMP
                )
                """
            )

            # --------------------------------------------------
            # Useful indexes
            # --------------------------------------------------

            cursor.execute(
                """
                CREATE INDEX IF NOT EXISTS
                idx_endurance_cycle
                ON endurance_data
                (
                    test_folder,
                    dut_no,
                    cycle_no
                )
                """
            )

            cursor.execute(
                """
                CREATE INDEX IF NOT EXISTS
                idx_endurance_csv
                ON endurance_data
                (
                    csv_written
                )
                """
            )

            cursor.execute(
                """
                CREATE INDEX IF NOT EXISTS
                idx_endurance_excel
                ON endurance_data
                (
                    excel_written
                )
                """
            )

            conn.commit()

        finally:

            conn.close()

    # ==========================================================
    # SQLITE CONNECTION
    # ==========================================================

    def _get_db_connection(self):

        return sqlite3.connect(
            self.db_path,
            timeout=30
        )

    # ==========================================================
    # SANITIZE NAME
    # ==========================================================

    def sanitize_name(
        self,
        name
    ):

        if name is None:

            name = "Unknown_Test"

        name = str(name)

        invalid = '<>:"/\\|?*'

        for char in invalid:

            name = name.replace(
                char,
                "_"
            )

        return name.strip()

    # ==========================================================
    # START TEST
    # ==========================================================

    def start_test(
        self,
        test_name
    ):

        self.test_name = self.sanitize_name(
            test_name
        )

        timestamp = datetime.now().strftime(
            "%Y-%m-%d_%H-%M-%S"
        )

        self.test_folder_name = (
            f"{self.test_name}_{timestamp}"
        )

        # ------------------------------------------------------
        # CSV ROOT
        # ------------------------------------------------------

        csv_root = os.path.join(
            self.base_folder,
            self.test_folder_name
        )

        os.makedirs(
            csv_root,
            exist_ok=True
        )

        # ------------------------------------------------------
        # Excel ROOT
        # ------------------------------------------------------

        excel_root = os.path.join(
            self.report_folder,
            "Endurance",
            self.test_folder_name
        )

        os.makedirs(
            excel_root,
            exist_ok=True
        )

        print()
        print(
            "=========================================="
        )

        print(
            "Data logging started"
        )

        print(
            f"CSV    : {csv_root}"
        )

        print(
            f"Report : {excel_root}"
        )

        print(
            f"SQLite : {self.db_path}"
        )

        print(
            "=========================================="
        )

    # ==========================================================
    # START CYCLE
    # ==========================================================

    def start_cycle(
        self,
        test_name,
        dut_no,
        cycle_no,
        serial_number="",
        user_name="",
        temperature_type="",
        initial_temperature="",
        source_type=""
    ):

        # ------------------------------------------------------
        # Automatically start test if necessary
        # ------------------------------------------------------

        if self.test_folder_name is None:

            self.start_test(
                test_name
            )

        key = (
            dut_no,
            cycle_no
        )

        # ======================================================
        # CSV
        # ======================================================

        dut_csv_folder = os.path.join(
            self.base_folder,
            self.test_folder_name,
            f"DUT{dut_no}"
        )

        os.makedirs(
            dut_csv_folder,
            exist_ok=True
        )

        csv_path = os.path.join(
            dut_csv_folder,
            f"Cycle_{cycle_no}.csv"
        )

        self.csv_paths[key] = csv_path

        # ------------------------------------------------------
        # Create CSV header
        # ------------------------------------------------------

        self._create_csv_file(
            csv_path
        )

        # ======================================================
        # EXCEL
        # ======================================================

        dut_excel_folder = os.path.join(
            self.report_folder,
            "Endurance",
            self.test_folder_name,
            f"DUT{dut_no}"
        )

        os.makedirs(
            dut_excel_folder,
            exist_ok=True
        )

        excel_path = os.path.join(
            dut_excel_folder,
            f"Cycle_{cycle_no}.xlsx"
        )

        # ======================================================
        # START TIME
        # ======================================================

        start_time = datetime.now()

        self.cycle_start_time[key] = (
            start_time
        )

        # ======================================================
        # CREATE EXCEL WORKBOOK
        # ======================================================

        workbook = Workbook()

        worksheet = workbook.active

        worksheet.title = (
            f"Cycle {cycle_no}"
        )

        self.excel_files[key] = {

            "workbook": workbook,

            "worksheet": worksheet,

            "path": excel_path,

            # --------------------------------------------------
            # Excel data rows
            # --------------------------------------------------

            "charging_rows": [],

            "discharging_rows": [],

            # --------------------------------------------------
            # SQLite IDs already written into Excel
            # --------------------------------------------------

            "charging_record_ids": [],

            "discharging_record_ids": [],

            # --------------------------------------------------
            # Excel section positions
            # --------------------------------------------------

            "charging_start_row": None,

            "charging_header_row": None,

            "discharging_start_row": None,

            "discharging_header_row": None,

            # --------------------------------------------------
            # Save state
            # --------------------------------------------------

            "dirty": True,

            "last_save_time": 0,

            "finished": False
        }

        # ------------------------------------------------------
        # Create Excel header
        # ------------------------------------------------------

        self._create_excel_header(
            key=key,
            serial_number=serial_number,
            dut_no=dut_no,
            start_time=start_time,
            user_name=user_name,
            temperature_type=temperature_type,
            initial_temperature=initial_temperature,
            source_type=source_type
        )

        # ------------------------------------------------------
        # Initial Excel save
        # ------------------------------------------------------

        try:

            workbook.save(
                excel_path
            )

            self.excel_files[key][
                "last_save_time"
            ] = time.time()

        except PermissionError as e:

            print(
                f"Excel initial save failed: {e}"
            )

        print(
            f"CSV started   : {csv_path}"
        )

        print(
            f"Excel started : {excel_path}"
        )

    # ==========================================================
    # CREATE CSV
    # ==========================================================

    def _create_csv_file(
        self,
        csv_path
    ):

        # ------------------------------------------------------
        # If file does not exist, create header
        # ------------------------------------------------------

        if not os.path.exists(
            csv_path
        ):

            with open(
                csv_path,
                "w",
                newline="",
                encoding="utf-8"
            ) as file:

                writer = csv.DictWriter(
                    file,
                    fieldnames=self._get_data_headers()
                )

                writer.writeheader()

        else:

            # --------------------------------------------------
            # Existing file:
            # Don't overwrite it.
            # --------------------------------------------------

            pass

    # ==========================================================
    # EXCEL HEADER
    # ==========================================================

    def _create_excel_header(
        self,
        key,
        serial_number,
        dut_no,
        start_time,
        user_name,
        temperature_type,
        initial_temperature,
        source_type
    ):

        info = self.excel_files[key]

        ws = info["worksheet"]

        # ------------------------------------------------------
        # Title
        # ------------------------------------------------------

        ws["A1"] = (
            "Endurance Test Report"
        )

        ws["A1"].font = Font(
            bold=True,
            size=16
        )

        ws.merge_cells(
            start_row=1,
            start_column=1,
            end_row=1,
            end_column=6
        )

        # ------------------------------------------------------
        # Information
        # ------------------------------------------------------

        headers = [

            "Serial Number",
            "DUT Slot",
            "Start Time",
            "User Name",
            "Temperature Type",
            "Initial Temperature(°C)",
            "Source Type"
        ]

        values = [

            serial_number,

            f"DUT{dut_no}",

            start_time.strftime(
                "%Y-%m-%d %H:%M:%S"
            ),

            user_name,

            temperature_type,

            initial_temperature,

            source_type
        ]

        # ------------------------------------------------------
        # Headers
        # ------------------------------------------------------

        for column, value in enumerate(
            headers,
            start=1
        ):

            cell = ws.cell(
                row=3,
                column=column
            )

            cell.value = value

            cell.font = Font(
                bold=True
            )

        # ------------------------------------------------------
        # Values
        # ------------------------------------------------------

        for column, value in enumerate(
            values,
            start=1
        ):

            ws.cell(
                row=4,
                column=column
            ).value = value

        # ------------------------------------------------------
        # Alignment
        # ------------------------------------------------------

        for row in ws.iter_rows(
            min_row=3,
            max_row=4,
            min_col=1,
            max_col=len(headers)
        ):

            for cell in row:

                cell.alignment = Alignment(
                    horizontal="center",
                    vertical="center"
                )

        # ======================================================
        # CHARGING SECTION
        # ======================================================

        row = 6

        ws.cell(
            row=row,
            column=1
        ).value = (
            "OBC Charging"
        )

        ws.cell(
            row=row,
            column=1
        ).font = Font(
            bold=True,
            size=14
        )

        info[
            "charging_start_row"
        ] = row

        row += 1

        info[
            "charging_header_row"
        ] = row

        self._write_data_headers(
            ws,
            row,
            self._get_data_headers()
        )

    # ==========================================================
    # DATA HEADERS
    # ==========================================================

    # def _get_data_headers(self):

    #     headers = [

    #         "Timestamp",
    #         "Time_sec",
    #         "Mode"
    #     ]

    #     # ------------------------------------------------------
    #     # CAN
    #     # ------------------------------------------------------

    #     for parameter in (
    #         self.COMMON_PARAMETERS
    #     ):

    #         headers.append(
    #             f"CAN_{parameter}"
    #         )

    #     # ------------------------------------------------------
    #     # Hardware
    #     # ------------------------------------------------------

    #     for parameter in (
    #         self.COMMON_PARAMETERS
    #     ):

    #         headers.append(
    #             f"HW_{parameter}"
    #         )

    #     return headers


    def _get_data_headers(self):

        headers = [
            "Timestamp",
            "Time_sec",
            "Mode"
        ]

        # ======================================================
        # CAN
        # ======================================================

        for parameter in self.COMMON_PARAMETERS:

            unit = self.PARAMETER_UNITS.get(
                parameter,
                ""
            )

            if unit:
                display_name = (
                    f"{parameter} ({unit})"
                )
            else:
                display_name = parameter

            headers.append(
                f"CAN_{display_name}"
            )

        # ======================================================
        # HARDWARE
        # ======================================================

        for parameter in self.COMMON_PARAMETERS:

            unit = self.PARAMETER_UNITS.get(
                parameter,
                ""
            )

            if unit:
                display_name = (
                    f"{parameter} ({unit})"
                )
            else:
                display_name = parameter

            headers.append(
                f"HW_{display_name}"
            )

        return headers

    def _get_display_parameter_name(
        self,
        parameter
        ):

        unit = self.PARAMETER_UNITS.get(
            parameter,
            ""
        )

        if unit:

            return (
                f"{parameter} ({unit})"
            )

        return parameter


    def _build_export_row(
        self,
        row
    ):

        export_row = {

            "Timestamp": row.get(
                "Timestamp",
                ""
            ),

            "Time_sec": row.get(
                "Time_sec",
                ""
            ),

            "Mode": row.get(
                "Mode",
                ""
            )
        }

        for parameter in self.COMMON_PARAMETERS:

            display_name = (
                self._get_display_parameter_name(
                    parameter
                )
            )

            export_row[
                f"CAN_{display_name}"
            ] = row.get(
                f"CAN_{parameter}",
                ""
            )

            export_row[
                f"HW_{display_name}"
            ] = row.get(
                f"HW_{parameter}",
                ""
            )

        return export_row
    
    # ==========================================================
    # WRITE EXCEL HEADERS
    # ==========================================================

    def _write_data_headers(
        self,
        worksheet,
        row,
        headers
    ):

        for column, header in enumerate(
            headers,
            start=1
        ):

            cell = worksheet.cell(
                row=row,
                column=column
            )

            cell.value = header

            cell.font = Font(
                bold=True
            )

            cell.alignment = Alignment(
                horizontal="center",
                vertical="center"
            )

    # ==========================================================
    # LOG DATA
    # ==========================================================

    # def log_data(
    #     self,
    #     dut_no,
    #     cycle_no,
    #     mode,
    #     can_values,
    #     hardware_values,
    #     timestamp=None,
    #     time_sec=None
    # ):

    #     key = (
    #         dut_no,
    #         cycle_no
    #     )

    #     # ------------------------------------------------------
    #     # Timestamp
    #     # ------------------------------------------------------

    #     if timestamp is None:

    #         timestamp = datetime.now()

    #     if isinstance(
    #         timestamp,
    #         datetime
    #     ):

    #         timestamp_text = (
    #             timestamp.strftime(
    #                 "%Y-%m-%d %H:%M:%S"
    #             )
    #         )

    #     else:

    #         timestamp_text = str(
    #             timestamp
    #         )

    #     # ------------------------------------------------------
    #     # Time
    #     # ------------------------------------------------------

    #     if time_sec is None:

    #         start = self.cycle_start_time.get(
    #             key,
    #             timestamp
    #         )

    #         if (
    #             isinstance(start, datetime)
    #             and
    #             isinstance(timestamp, datetime)
    #         ):

    #             time_sec = (
    #                 timestamp - start
    #             ).total_seconds()

    #         else:

    #             time_sec = 0

    #     # ======================================================
    #     # BUILD ROW
    #     # ======================================================

    #     row = {

    #         "Timestamp": timestamp_text,

    #         "Time_sec": round(
    #             float(time_sec),
    #             2
    #         ),

    #         "Mode": mode
    #     }

    #     # ------------------------------------------------------
    #     # CAN values
    #     # ------------------------------------------------------

    #     if can_values is None:

    #         can_values = {}

    #     for parameter in (
    #         self.COMMON_PARAMETERS
    #     ):

    #         row[
    #             f"CAN_{parameter}"
    #         ] = can_values.get(
    #             parameter,
    #             ""
    #         )

    #     # ------------------------------------------------------
    #     # Hardware values
    #     # ------------------------------------------------------

    #     if hardware_values is None:

    #         hardware_values = {}

    #     for parameter in (
    #         self.COMMON_PARAMETERS
    #     ):

    #         row[
    #             f"HW_{parameter}"
    #         ] = hardware_values.get(
    #             parameter,
    #             ""
    #         )

    #     # ======================================================
    #     # PUT INTO QUEUE
    #     # ======================================================

    #     self.log_queue.put({

    #         "dut_no": dut_no,

    #         "cycle_no": cycle_no,

    #         "row": copy.deepcopy(
    #             row
    #         )
    #     })


    def log_data(
        self,
        dut_no,
        cycle_no,
        mode,
        can_values,
        hardware_values,
        timestamp=None,
        time_sec=None
    ):

        key = (
            dut_no,
            cycle_no
        )

        # ======================================================
        # TIMESTAMP
        # ======================================================

        if timestamp is None:
            timestamp = datetime.now()

        # Keep datetime available for elapsed-time calculation
        timestamp_dt = timestamp if isinstance(
            timestamp,
            datetime
        ) else datetime.now()

        timestamp_text = timestamp_dt.strftime(
            "%Y-%m-%d %H:%M:%S"
        )

        # ======================================================
        # TIME_SEC
        # ======================================================

        # IMPORTANT:
        # Calculate from the START OF THE ENTIRE CYCLE.
        #
        # Do not depend on cycle_elapsed coming from
        # wait_seconds().
        # ======================================================

        start_time = self.cycle_start_time.get(
            key
        )

        if start_time is not None:

            calculated_time_sec = (
                timestamp_dt - start_time
            ).total_seconds()

            time_sec = round(
                calculated_time_sec,
                2
            )

        elif time_sec is None:

            time_sec = 0

        else:

            time_sec = round(
                float(time_sec),
                2
            )

        # ======================================================
        # BUILD ROW
        # ======================================================

        row = {

            "Timestamp": timestamp_text,

            "Time_sec": time_sec,

            "Mode": mode
        }

        # ======================================================
        # CAN
        # ======================================================

        if can_values is None:
            can_values = {}

        for parameter in self.COMMON_PARAMETERS:

            row[
                f"CAN_{parameter}"
            ] = can_values.get(
                parameter,
                ""
            )

        # ======================================================
        # HARDWARE
        # ======================================================

        if hardware_values is None:
            hardware_values = {}

        for parameter in self.COMMON_PARAMETERS:

            row[
                f"HW_{parameter}"
            ] = hardware_values.get(
                parameter,
                ""
            )

        # ======================================================
        # QUEUE
        # ======================================================

        self.log_queue.put({

            "dut_no": dut_no,

            "cycle_no": cycle_no,

            "row": copy.deepcopy(
                row
            )
        })

    # ==========================================================
    # LOGGER WORKER
    # ==========================================================

    def _logger_worker(self):

        print(
            "Logger worker running."
        )

        while (
            self.running
            or not self.log_queue.empty()
        ):

            # --------------------------------------------------
            # Process queued measurements
            # --------------------------------------------------

            try:

                item = self.log_queue.get(
                    timeout=0.5
                )

            except queue.Empty:

                # ----------------------------------------------
                # Retry pending exports
                # ----------------------------------------------

                self._retry_pending_exports()

                continue

            try:

                self._process_measurement(
                    item
                )

            except Exception as e:

                print(
                    f"Logger worker error: {e}"
                )

                # ----------------------------------------------
                # Put item back into queue
                # ----------------------------------------------

                self.log_queue.put(
                    item
                )

                time.sleep(
                    self.RETRY_INTERVAL
                )

            finally:

                self.log_queue.task_done()

    # ==========================================================
    # PROCESS ONE MEASUREMENT
    # ==========================================================

    def _process_measurement(
        self,
        item
    ):

        dut_no = item["dut_no"]

        cycle_no = item["cycle_no"]

        row = item["row"]

        # ======================================================
        # STEP 1
        # SQLITE
        # ======================================================

        record_id = self._insert_sqlite(
            dut_no,
            cycle_no,
            row
        )

        # ======================================================
        # STEP 2
        # CSV
        # ======================================================

        try:

            self._write_csv(
                dut_no,
                cycle_no,
                row
            )

            self._mark_csv_written(
                record_id
            )

        except (
            PermissionError,
            OSError
        ) as e:

            print(
                f"CSV locked/unavailable "
                f"DUT{dut_no} "
                f"Cycle {cycle_no}: "
                f"{e}"
            )

            # --------------------------------------------------
            # DO NOT retry SQLite.
            #
            # SQLite record already exists.
            # --------------------------------------------------

        # ======================================================
        # STEP 3
        # EXCEL
        # ======================================================

        try:

            self._write_excel(
                dut_no,
                cycle_no,
                row,
                record_id
            )

        except (
            PermissionError,
            OSError
        ) as e:

            print(
                f"Excel locked/unavailable "
                f"DUT{dut_no} "
                f"Cycle {cycle_no}: "
                f"{e}"
            )

    # ==========================================================
    # SQLITE INSERT
    # ==========================================================

    def _insert_sqlite(
        self,
        dut_no,
        cycle_no,
        row
    ):

        conn = self._get_db_connection()

        try:

            cursor = conn.cursor()

            # --------------------------------------------------
            # Separate CAN / HW data
            # --------------------------------------------------

            can_values = {}

            hardware_values = {}

            for parameter in (
                self.COMMON_PARAMETERS
            ):

                can_values[
                    parameter
                ] = row.get(
                    f"CAN_{parameter}",
                    ""
                )

                hardware_values[
                    parameter
                ] = row.get(
                    f"HW_{parameter}",
                    ""
                )

            # --------------------------------------------------
            # INSERT
            # --------------------------------------------------

            cursor.execute(
                """
                INSERT INTO endurance_data
                (
                    test_name,
                    test_folder,
                    dut_no,
                    cycle_no,
                    timestamp,
                    time_sec,
                    mode,
                    can_values,
                    hardware_values,
                    csv_written,
                    excel_written
                )
                VALUES
                (
                    ?, ?, ?, ?, ?, ?,
                    ?, ?, ?, 0, 0
                )
                """,
                (

                    self.test_name,

                    self.test_folder_name,

                    dut_no,

                    cycle_no,

                    row["Timestamp"],

                    row["Time_sec"],

                    row["Mode"],

                    json.dumps(
                        can_values
                    ),

                    json.dumps(
                        hardware_values
                    )
                )
            )

            record_id = (
                cursor.lastrowid
            )

            conn.commit()

            return record_id

        finally:

            conn.close()

    # ==========================================================
    # CSV WRITE
    # ==========================================================

    # def _write_csv(
    #     self,
    #     dut_no,
    #     cycle_no,
    #     row
    # ):

    #     key = (
    #         dut_no,
    #         cycle_no
    #     )

    #     if key not in self.csv_paths:

    #         raise RuntimeError(
    #             f"CSV path not found "
    #             f"for DUT{dut_no} "
    #             f"Cycle {cycle_no}"
    #         )

    #     csv_path = self.csv_paths[key]

    #     # ------------------------------------------------------
    #     # IMPORTANT:
    #     #
    #     # Open file only for the duration of this write.
    #     #
    #     # This makes file locking much easier to detect.
    #     # ------------------------------------------------------

    #     with open(
    #         csv_path,
    #         "a",
    #         newline="",
    #         encoding="utf-8"
    #     ) as file:

    #         writer = csv.DictWriter(
    #             file,
    #             fieldnames=self._get_data_headers()
    #         )

    #         writer.writerow(
    #             row
    #         )

    #         file.flush()

    def _write_csv(
        self,
        dut_no,
        cycle_no,
        row
    ):

        key = (
            dut_no,
            cycle_no
        )

        if key not in self.csv_paths:

            raise RuntimeError(
                f"CSV path not found "
                f"for DUT{dut_no} "
                f"Cycle {cycle_no}"
            )

        csv_path = self.csv_paths[key]

        export_row = self._build_export_row(
            row
        )

        with open(
            csv_path,
            "a",
            newline="",
            encoding="utf-8"
        ) as file:

            writer = csv.DictWriter(
                file,
                fieldnames=self._get_data_headers()
            )

            writer.writerow(
                export_row
            )

            file.flush()

    # ==========================================================
    # MARK CSV WRITTEN
    # ==========================================================

    def _mark_csv_written(
        self,
        record_id
    ):

        conn = self._get_db_connection()

        try:

            conn.execute(
                """
                UPDATE endurance_data

                SET csv_written = 1

                WHERE id = ?
                """,
                (
                    record_id,
                )
            )

            conn.commit()

        finally:

            conn.close()

    # ==========================================================
    # MARK EXCEL WRITTEN
    # ==========================================================

    def _mark_excel_written(
        self,
        record_ids
    ):

        if not record_ids:

            return

        conn = self._get_db_connection()

        try:

            cursor = conn.cursor()

            cursor.executemany(
                """
                UPDATE endurance_data

                SET excel_written = 1

                WHERE id = ?
                """,
                [
                    (
                        record_id,
                    )
                    for record_id in record_ids
                ]
            )

            conn.commit()

        finally:

            conn.close()

    # ==========================================================
    # WRITE EXCEL
    # ==========================================================

    def _write_excel(
        self,
        dut_no,
        cycle_no,
        row,
        record_id
    ):

        key = (
            dut_no,
            cycle_no
        )

        if key not in self.excel_files:

            raise RuntimeError(
                f"Excel not initialized "
                f"for DUT{dut_no} "
                f"Cycle {cycle_no}"
            )

        info = self.excel_files[key]

        ws = info["worksheet"]

        mode_lower = str(
            row["Mode"]
        ).lower()

        # ======================================================
        # CHARGING
        # ======================================================

        if mode_lower in (
            "charge",
            "charging"
        ):

            section = "charging"

        # ======================================================
        # DISCHARGING
        # ======================================================

        elif mode_lower in (
            "discharge",
            "discharging"
        ):

            section = "discharging"

        # ======================================================
        # REST
        # ======================================================

        else:

            # --------------------------------------------------
            # REST is stored in SQLite and CSV.
            #
            # It is intentionally not put into the
            # Charging / Discharging Excel sections.
            # --------------------------------------------------

            return

        # ======================================================
        # CREATE DISCHARGE SECTION
        # ======================================================

        if (
            section == "discharging"
            and
            info[
                "discharging_header_row"
            ] is None
        ):

            self._create_discharging_section(
                key
            )

        # ======================================================
        # SELECT SECTION
        # ======================================================

        if section == "charging":

            header_row = (
                info[
                    "charging_header_row"
                ]
            )

            data_rows = (
                info[
                    "charging_rows"
                ]
            )

            record_ids = (
                info[
                    "charging_record_ids"
                ]
            )

        else:

            header_row = (
                info[
                    "discharging_header_row"
                ]
            )

            data_rows = (
                info[
                    "discharging_rows"
                ]
            )

            record_ids = (
                info[
                    "discharging_record_ids"
                ]
            )

        # ======================================================
        # EXCEL ROW
        # ======================================================

        excel_row = (
            header_row
            + 1
            + len(data_rows)
        )

        # ======================================================
        # WRITE VALUES
        # ======================================================

        # for column, header in enumerate(
        #     self._get_data_headers(),
        #     start=1
        # ):

        #     ws.cell(
        #         row=excel_row,
        #         column=column
        #     ).value = row.get(
        #         header,
        #         ""
        #     )

        export_row = self._build_export_row(
            row
        )

        for column, header in enumerate(
            self._get_data_headers(),
            start=1
        ):

            ws.cell(
                row=excel_row,
                column=column
            ).value = export_row.get(
                header,
                ""
            )

        # ======================================================
        # TRACK ROW
        # ======================================================

        data_rows.append(
            excel_row
        )

        record_ids.append(
            record_id
        )

        info["dirty"] = True

        # ======================================================
        # PERIODIC SAVE
        # ======================================================

        now = time.time()

        if (
            now
            -
            info["last_save_time"]
            >=
            self.EXCEL_SAVE_INTERVAL
        ):

            self._save_excel(
                key
            )

    # ==========================================================
    # CREATE DISCHARGE SECTION
    # ==========================================================

    def _create_discharging_section(
        self,
        key
    ):

        info = self.excel_files[key]

        ws = info["worksheet"]

        row = (
            ws.max_row + 3
        )

        # ------------------------------------------------------
        # Heading
        # ------------------------------------------------------

        ws.cell(
            row=row,
            column=1
        ).value = (
            "DCDC Discharging"
        )

        ws.cell(
            row=row,
            column=1
        ).font = Font(
            bold=True,
            size=14
        )

        info[
            "discharging_start_row"
        ] = row

        row += 1

        # ------------------------------------------------------
        # Header
        # ------------------------------------------------------

        info[
            "discharging_header_row"
        ] = row

        self._write_data_headers(
            ws,
            row,
            self._get_data_headers()
        )

    # ==========================================================
    # SAVE EXCEL
    # ==========================================================

    def _save_excel(
        self,
        key
    ):

        if key not in self.excel_files:

            return False

        info = self.excel_files[key]

        if not info["dirty"]:

            return True

        try:

            # --------------------------------------------------
            # Formatting
            # --------------------------------------------------

            self._adjust_column_widths(
                info["worksheet"]
            )

            # --------------------------------------------------
            # Save workbook
            # --------------------------------------------------

            info["workbook"].save(
                info["path"]
            )

            # --------------------------------------------------
            # Mark corresponding SQLite records
            # --------------------------------------------------

            all_record_ids = (
                info[
                    "charging_record_ids"
                ]
                +
                info[
                    "discharging_record_ids"
                ]
            )

            self._mark_excel_written(
                all_record_ids
            )

            # --------------------------------------------------
            # Clear tracking
            #
            # The rows remain in Excel, but the IDs no longer
            # need to be exported again.
            # --------------------------------------------------

            info[
                "charging_record_ids"
            ].clear()

            info[
                "discharging_record_ids"
            ].clear()

            info["dirty"] = False

            info[
                "last_save_time"
            ] = time.time()

            return True

        except (
            PermissionError,
            OSError
        ):

            # --------------------------------------------------
            # Excel is probably locked.
            #
            # DO NOT delete anything.
            #
            # Workbook remains in memory.
            # SQLite remains safe.
            # Worker will retry later.
            # --------------------------------------------------

            return False

    # ==========================================================
    # RETRY PENDING EXPORTS
    # ==========================================================

    def _retry_pending_exports(self):

        # ------------------------------------------------------
        # Retry Excel files that are dirty
        # ------------------------------------------------------

        for key in list(
            self.excel_files.keys()
        ):

            info = self.excel_files[key]

            if info["dirty"]:

                self._save_excel(
                    key
                )

        # ------------------------------------------------------
        # Retry CSV records that failed earlier
        # ------------------------------------------------------

        self._retry_pending_csv()

    # ==========================================================
    # RETRY CSV
    # ==========================================================

    def _retry_pending_csv(self):

        if self.test_folder_name is None:

            return

        conn = self._get_db_connection()

        try:

            cursor = conn.cursor()

            cursor.execute(
                """
                SELECT

                    id,
                    dut_no,
                    cycle_no,
                    timestamp,
                    time_sec,
                    mode,
                    can_values,
                    hardware_values

                FROM endurance_data

                WHERE

                    test_folder = ?

                    AND csv_written = 0

                ORDER BY id
                """,
                (
                    self.test_folder_name,
                )
            )

            records = cursor.fetchall()

        finally:

            conn.close()

        # ------------------------------------------------------
        # Export records
        # ------------------------------------------------------

        for record in records:

            (
                record_id,
                dut_no,
                cycle_no,
                timestamp,
                time_sec,
                mode,
                can_json,
                hardware_json
            ) = record

            key = (
                dut_no,
                cycle_no
            )

            # --------------------------------------------------
            # Cycle may already have been closed.
            #
            # Reconstruct CSV path.
            # --------------------------------------------------

            if key not in self.csv_paths:

                csv_path = os.path.join(

                    self.base_folder,

                    self.test_folder_name,

                    f"DUT{dut_no}",

                    f"Cycle_{cycle_no}.csv"
                )

            else:

                csv_path = (
                    self.csv_paths[key]
                )

            row = {

                "Timestamp": timestamp,

                "Time_sec": time_sec,

                "Mode": mode
            }

            can_values = json.loads(
                can_json or "{}"
            )

            hardware_values = json.loads(
                hardware_json or "{}"
            )

            for parameter in (
                self.COMMON_PARAMETERS
            ):

                row[
                    f"CAN_{parameter}"
                ] = can_values.get(
                    parameter,
                    ""
                )

                row[
                    f"HW_{parameter}"
                ] = hardware_values.get(
                    parameter,
                    ""
                )

            try:

                # --------------------------------------------------
                # Make sure file exists/header exists
                # --------------------------------------------------

                self._create_csv_file(
                    csv_path
                )

                with open(
                    csv_path,
                    "a",
                    newline="",
                    encoding="utf-8"
                ) as file:

                    writer = csv.DictWriter(
                        file,
                        fieldnames=self._get_data_headers()
                    )

                    writer.writerow(
                        row
                    )

                    file.flush()

                # --------------------------------------------------
                # Mark record exported
                # --------------------------------------------------

                self._mark_csv_written(
                    record_id
                )

            except (
                PermissionError,
                OSError
            ):

                # --------------------------------------------------
                # Still locked.
                #
                # Stop retrying this round.
                # --------------------------------------------------

                break

    # ==========================================================
    # FINISH CYCLE
    # ==========================================================

    def finish_cycle(
        self,
        dut_no,
        cycle_no
    ):

        key = (
            dut_no,
            cycle_no
        )

        print(
            f"\nFinishing "
            f"DUT{dut_no} "
            f"Cycle {cycle_no}..."
        )

        # ======================================================
        # WAIT FOR QUEUE
        # ======================================================

        self.log_queue.join()

        # ======================================================
        # FINAL EXCEL SAVE
        # ======================================================

        if key in self.excel_files:

            info = self.excel_files[key]

            # --------------------------------------------------
            # Average rows
            # --------------------------------------------------

            ws = info["worksheet"]

            if info[
                "charging_rows"
            ]:

                self._write_average_row(
                    ws,
                    info[
                        "charging_rows"
                    ],
                    "charging"
                )

            if info[
                "discharging_rows"
            ]:

                self._write_average_row(
                    ws,
                    info[
                        "discharging_rows"
                    ],
                    "discharging"
                )

            # ======================================================
            # IMPORTANT FIX
            # ======================================================
            info["dirty"] = True

            # --------------------------------------------------
            # Format
            # --------------------------------------------------

            self._format_excel(
                ws
            )

            info["finished"] = True

            # --------------------------------------------------
            # Try save
            # --------------------------------------------------

            success = self._save_excel(
                key
            )

            if success:

                print(
                    f"Excel saved: "
                    f"DUT{dut_no} "
                    f"Cycle {cycle_no}"
                )

                # ----------------------------------------------
                # Remove Excel object
                # ----------------------------------------------

                del self.excel_files[key]

            else:

                # ------------------------------------------------
                # IMPORTANT
                #
                # Excel is locked.
                #
                # We DO NOT delete the workbook.
                #
                # Worker will continue retrying.
                # ------------------------------------------------

                print(
                    f"Excel is locked. "
                    f"DUT{dut_no} "
                    f"Cycle {cycle_no} "
                    f"will be retried in background."
                )

        # ======================================================
        # CSV
        # ======================================================

        # ------------------------------------------------------
        # We do not keep CSV handles open.
        #
        # Therefore there is nothing to close here.
        # ------------------------------------------------------

        # ======================================================
        # REMOVE START TIME
        # ======================================================

        self.cycle_start_time.pop(
            key,
            None
        )

        print(
            f"Cycle processing completed: "
            f"DUT{dut_no} "
            f"Cycle {cycle_no}"
        )

    # ==========================================================
    # WRITE AVERAGE
    # ==========================================================

    # def _write_average_row(
    #     self,
    #     worksheet,
    #     data_rows,
    #     section
    # ):

    #     if not data_rows:

    #         return

    #     average_row = (
    #         max(data_rows) + 1
    #     )

    #     worksheet.cell(
    #         row=average_row,
    #         column=1
    #     ).value = "Average"

    #     worksheet.cell(
    #         row=average_row,
    #         column=1
    #     ).font = Font(
    #         bold=True
    #     )

    #     headers = (
    #         self._get_data_headers()
    #     )

    #     for column, header in enumerate(
    #         headers,
    #         start=1
    #     ):

    #         if header in (
    #             "Timestamp",
    #             "Mode"
    #         ):

    #             continue

    #         values = []

    #         for row in data_rows:

    #             value = worksheet.cell(
    #                 row=row,
    #                 column=column
    #             ).value

    #             try:

    #                 if (
    #                     value is not None
    #                     and value != ""
    #                 ):

    #                     values.append(
    #                         float(value)
    #                     )

    #             except (
    #                 ValueError,
    #                 TypeError
    #             ):

    #                 pass

    #         if values:

    #             average = (
    #                 sum(values)
    #                 /
    #                 len(values)
    #             )

    #             worksheet.cell(
    #                 row=average_row,
    #                 column=column
    #             ).value = round(
    #                 average,
    #                 4
    #             )

    def _write_average_row(
        self,
        worksheet,
        data_rows,
        section
    ):

        if not data_rows:
            return

        average_row = (
            max(data_rows) + 1
        )

        headers = (
            self._get_data_headers()
        )

        # ======================================================
        # CALCULATE AVERAGES
        # ======================================================

        for column, header in enumerate(
            headers,
            start=1
        ):

            if header in (
                "Timestamp",
                "Mode",
                "Time_sec"
            ):

                continue

            values = []

            for row in data_rows:

                value = worksheet.cell(
                    row=row,
                    column=column
                ).value

                try:

                    if (
                        value is not None
                        and value != ""
                    ):

                        values.append(
                            float(value)
                        )

                except (
                    ValueError,
                    TypeError
                ):

                    pass

            if values:

                average = (
                    sum(values)
                    /
                    len(values)
                )

                worksheet.cell(
                    row=average_row,
                    column=column
                ).value = round(
                    average,
                    4
                )

        # ======================================================
        # AVERAGE LABEL
        # ======================================================

        worksheet.cell(
            row=average_row,
            column=1
        ).value = "Average"

        # ======================================================
        # AVERAGE ROW STYLE
        # ======================================================

        if section == "charging":

            average_fill = PatternFill(
                fill_type="solid",
                fgColor="FFF2CC"     # Light yellow
            )

        else:

            average_fill = PatternFill(
                fill_type="solid",
                fgColor="DDEBF7"     # Light blue
            )

        average_font = Font(
            bold=True
        )

        average_alignment = Alignment(
            horizontal="center",
            vertical="center"
        )

        # ======================================================
        # APPLY STYLE TO ENTIRE ROW
        # ======================================================

        for column in range(
            1,
            len(headers) + 1
        ):

            cell = worksheet.cell(
                row=average_row,
                column=column
            )

            cell.fill = average_fill
            cell.font = average_font
            cell.alignment = average_alignment

    # ==========================================================
    # FORMAT EXCEL
    # ==========================================================

    def _format_excel(
        self,
        worksheet
    ):

        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin")
        )

        for row in worksheet.iter_rows():

            for cell in row:

                cell.alignment = Alignment(
                    vertical="center",
                    horizontal="center"
                )

                if cell.value is not None:

                    cell.border = (
                        thin_border
                    )

        # ------------------------------------------------------
        # Section headings
        # ------------------------------------------------------

        for row in range(
            1,
            worksheet.max_row + 1
        ):

            value = worksheet.cell(
                row=row,
                column=1
            ).value

            if value in (
                "OBC Charging",
                "DCDC Discharging"
            ):

                worksheet.cell(
                    row=row,
                    column=1
                ).font = Font(
                    bold=True,
                    size=14
                )

        worksheet.freeze_panes = "A6"

    # ==========================================================
    # COLUMN WIDTH
    # ==========================================================

    def _adjust_column_widths(
        self,
        worksheet
    ):

        for column_cells in (
            worksheet.columns
        ):

            max_length = 0

            column_letter = (
                get_column_letter(
                    column_cells[0].column
                )
            )

            for cell in column_cells:

                try:

                    length = len(
                        str(cell.value)
                    )

                    if length > max_length:

                        max_length = length

                except Exception:

                    pass

            worksheet.column_dimensions[
                column_letter
            ].width = min(
                max(
                    max_length + 2,
                    12
                ),
                30
            )

    # ==========================================================
    # CLOSE ALL
    # ==========================================================

    def close_all(self):

        print()
        print(
            "Stopping endurance data logger..."
        )

        # ======================================================
        # WAIT FOR QUEUED DATA
        # ======================================================

        self.log_queue.join()

        # ======================================================
        # TRY FINAL EXCEL SAVES
        # ======================================================

        for key in list(
            self.excel_files.keys()
        ):

            info = self.excel_files[key]

            try:

                self._format_excel(
                    info["worksheet"]
                )

                success = self._save_excel(
                    key
                )

                if not success:

                    print(
                        f"Excel still locked: "
                        f"{info['path']}"
                    )

            except Exception as e:

                print(
                    f"Excel close error: {e}"
                )

        # ======================================================
        # STOP WORKER
        # ======================================================

        self.running = False

        self.logger_thread.join(
            timeout=10
        )

        # ======================================================
        # RETRY PENDING CSV
        # ======================================================

        self._retry_pending_csv()

        # ======================================================
        # FINAL STATUS
        # ======================================================

        print(
            "Endurance data logger stopped."
        )

    # ==========================================================
    # GET PENDING COUNTS
    # ==========================================================

    def get_pending_counts(self):

        conn = self._get_db_connection()

        try:

            cursor = conn.cursor()

            cursor.execute(
                """
                SELECT
                    COUNT(
                        CASE
                            WHEN csv_written = 0
                            THEN 1
                        END
                    ),

                    COUNT(
                        CASE
                            WHEN excel_written = 0
                            THEN 1
                        END
                    )

                FROM endurance_data

                WHERE test_folder = ?
                """,
                (
                    self.test_folder_name,
                )
            )

            result = cursor.fetchone()

            return {
                "csv_pending": result[0],
                "excel_pending": result[1]
            }

        finally:

            conn.close()