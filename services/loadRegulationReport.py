
import os
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import (
    Font,
    Alignment,
    Border,
    Side,
    PatternFill
)


class LoadRegulationReport:

    # ==========================================================
    # INITIALIZATION
    # ==========================================================

    def __init__(
        self,
        base_folder="CSV_Logs",
        report_folder="Report",
        database_folder="Data"
    ):

        self.base_folder = base_folder
        self.report_folder = report_folder
        self.database_folder = database_folder

        self.workbook = None
        self.worksheet = None
        self.report_path = None

        self.obc_sections = []
        self.hpdcdc_sections = []

    # ==========================================================
    # START REPORT
    # ==========================================================

    def start_report(
        self,
        test_name="Load_Regulation",
        serial_number="",
        dut_no="",
        start_time=None,
        user_name="",
        temperature_type="",
        initial_temperature="",
        source_type=""
    ):

        if start_time is None:
            start_time = datetime.now()

        # ======================================================
        # PROJECT FOLDER
        # ======================================================

        project_folder = os.path.dirname(
            os.path.dirname(
                os.path.abspath(__file__)
            )
        )

        # ======================================================
        # REPORT FOLDER
        # ======================================================

        report_folder = os.path.join(
            project_folder,
            self.report_folder,
            "Load Regulation"
        )

        os.makedirs(
            report_folder,
            exist_ok=True
        )

        # ======================================================
        # FILE NAME
        # ======================================================

        timestamp = datetime.now().strftime(
            "%Y-%m-%d_%H-%M-%S"
        )

        file_name = (
            f"{test_name}_{timestamp}.xlsx"
        )

        # ======================================================
        # FULL REPORT PATH
        # ======================================================

        self.report_path = os.path.join(
            report_folder,
            file_name
        )

        # ======================================================
        # CREATE WORKBOOK
        # ======================================================

        self.workbook = Workbook()

        self.worksheet = self.workbook.active

        self.worksheet.title = "Load Regulation"

        # Reset sections
        self.obc_sections = []
        self.hpdcdc_sections = []

        # ======================================================
        # CREATE REPORT
        # ======================================================

        self._create_report(
            serial_number=serial_number,
            dut_no=dut_no,
            start_time=start_time,
            user_name=user_name,
            temperature_type=temperature_type,
            initial_temperature=initial_temperature,
            source_type=source_type
        )

        # ======================================================
        # FORMAT
        # ======================================================

        self._format_excel()

        # ======================================================
        # SAVE
        # ======================================================

        self.workbook.save(
            self.report_path
        )

        relative_path = os.path.relpath(
            self.report_path,
            project_folder
        )

        print(
            f"REPORT GENERATED: {relative_path}"
        )

        return self.report_path

    # ==========================================================
    # CREATE EXCEL HEADER
    # ==========================================================

    def _create_excel_header(
        self,
        serial_number,
        dut_no,
        start_time,
        user_name,
        temperature_type,
        initial_temperature,
        source_type
    ):

        ws = self.worksheet

        # ======================================================
        # TITLE
        # ======================================================

        ws.merge_cells(
            start_row=1,
            start_column=1,
            end_row=1,
            end_column=13
        )

        ws["A1"] = "Load Regulation Test Report"

        ws["A1"].font = Font(
            bold=True,
            size=16
        )

        ws["A1"].alignment = Alignment(
            horizontal="center",
            vertical="center"
        )

        # ======================================================
        # INFORMATION HEADERS
        # ======================================================

        headers = [
            "Serial Number",
            "DUT Slot",
            "Start Time",
            "User Name",
            "Temperature Type",
            "Initial Temperature(°C)",
            "Source Type"
        ]

        # ======================================================
        # VALUES
        # ======================================================

        if isinstance(start_time, datetime):

            formatted_start_time = (
                start_time.strftime(
                    "%Y-%m-%d %H:%M:%S"
                )
            )

        else:

            formatted_start_time = str(
                start_time
            )

        values = [
            serial_number,
            f"DUT{dut_no}" if dut_no else "",
            formatted_start_time,
            user_name,
            temperature_type,
            initial_temperature,
            source_type
        ]

        # ======================================================
        # HEADER ROW
        # ======================================================

        for column, value in enumerate(
            headers,
            start=1
        ):

            cell = ws.cell(
                row=3,
                column=column
            )

            cell.value = value

            cell.font = Font(
                bold=True,
                size=9
            )

            cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
                wrap_text=True
            )

        # ======================================================
        # VALUE ROW
        # ======================================================

        for column, value in enumerate(
            values,
            start=1
        ):

            cell = ws.cell(
                row=4,
                column=column
            )

            cell.value = value

            cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
                wrap_text=True
            )

    # ==========================================================
    # CREATE COMPLETE REPORT
    # ==========================================================

    def _create_report(
        self,
        serial_number,
        dut_no,
        start_time,
        user_name,
        temperature_type,
        initial_temperature,
        source_type
    ):

        self._create_excel_header(
            serial_number=serial_number,
            dut_no=dut_no,
            start_time=start_time,
            user_name=user_name,
            temperature_type=temperature_type,
            initial_temperature=initial_temperature,
            source_type=source_type
        )

        # ======================================================
        # OBC - 84V
        # ======================================================

        row = 6

        row = self._create_obc_section(
            row=row,
            hv_voltage="84Vdc",
            hv_current="25.6A"
        )

        # ======================================================
        # OBC - 100.8V
        # ======================================================

        row += 2

        row = self._create_obc_section(
            row=row,
            hv_voltage="100.8Vdc",
            hv_current="21.3A"
        )

        # ======================================================
        # OBC - 118V
        # ======================================================

        row += 2

        row = self._create_obc_section(
            row=row,
            hv_voltage="118Vdc",
            hv_current="18.2A"
        )

        # ======================================================
        # HP DCDC - 84V
        # ======================================================

        row += 3

        row = self._create_hpdcdc_section(
            row=row,
            hv_voltage="84Vdc"
        )

        # ======================================================
        # HP DCDC - 100.8V
        # ======================================================

        row += 2

        row = self._create_hpdcdc_section(
            row=row,
            hv_voltage="100.8Vdc"
        )

        # ======================================================
        # HP DCDC - 118V
        # ======================================================

        row += 2

        row = self._create_hpdcdc_section(
            row=row,
            hv_voltage="118Vdc"
        )

    # ==========================================================
    # CREATE OBC SECTION
    # ==========================================================

    def _create_obc_section(
        self,
        row,
        hv_voltage,
        hv_current
    ):

        ws = self.worksheet

        # ======================================================
        # STYLES
        # ======================================================

        thin = Side(
            style="thin",
            color="000000"
        )

        border = Border(
            left=thin,
            right=thin,
            top=thin,
            bottom=thin
        )

        green_fill = PatternFill(
            fill_type="solid",
            fgColor="D9EAD3"
        )

        center = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True
        )

        # ======================================================
        # SECTION INFORMATION
        # ======================================================

        section = {
            "hv_voltage": hv_voltage,
            "hv_current": hv_current,
            "title_row": row,
            "condition_row": row + 1,
            "header_row": row + 2,
            "data_start": row + 3,
            "data_end": row + 5
        }

        self.obc_sections.append(section)

        # ======================================================
        # TITLE
        # ======================================================

        # CAN A:F
        ws.merge_cells(
            start_row=row,
            start_column=1,
            end_row=row,
            end_column=6
        )

        # Power Analyser H:M
        ws.merge_cells(
            start_row=row,
            start_column=8,
            end_row=row,
            end_column=13
        )

        can_title = ws.cell(
            row=row,
            column=1
        )

        can_title.value = (
            "OBC Load Regulation - CAN Data"
        )

        can_title.font = Font(
            bold=True,
            size=10
        )

        can_title.fill = green_fill
        can_title.alignment = center

        pa_title = ws.cell(
            row=row,
            column=8
        )

        pa_title.value = (
            "OBC Load Regulation - "
            "Power Analyser Data"
        )

        pa_title.font = Font(
            bold=True,
            size=10
        )

        pa_title.fill = green_fill
        pa_title.alignment = center

        # Borders for merged areas
        for col in range(1, 7):

            cell = ws.cell(
                row=row,
                column=col
            )

            cell.border = border
            cell.fill = green_fill

        for col in range(8, 14):

            cell = ws.cell(
                row=row,
                column=col
            )

            cell.border = border
            cell.fill = green_fill

        # ======================================================
        # TEST CONDITION
        # ======================================================

        row += 1

        ws.merge_cells(
            start_row=row,
            start_column=1,
            end_row=row,
            end_column=6
        )

        ws.merge_cells(
            start_row=row,
            start_column=8,
            end_row=row,
            end_column=13
        )

        condition = (
            f"Set HV Voltage: {hv_voltage}; "
            f"HV Current: {hv_current}"
        )

        can_condition = ws.cell(
            row=row,
            column=1
        )

        can_condition.value = condition
        can_condition.font = Font(
            bold=True,
            size=9
        )
        can_condition.alignment = center

        pa_condition = ws.cell(
            row=row,
            column=8
        )

        pa_condition.value = condition
        pa_condition.font = Font(
            bold=True,
            size=9
        )
        pa_condition.alignment = center

        # ======================================================
        # COLUMN HEADERS
        # ======================================================

        row += 1

        can_headers = [
            "HV Load(%)",
            "Set HV Current(A)",
            "Input AC Current CAN(A)",
            "Input Power CAN(W)",
            "Output OBC Voltage CAN(V)",
            "Load Regulation(%)"
        ]

        pa_headers = [
            "HV Load(%)",
            "Set HV Current(A)",
            "Input AC Voltage Power Analyser(V)",
            "Input AC Current Power Analyser(A)",
            "OBC Output Voltage Power Analyser(V)",
            "Load Regulation(%)"
        ]

        # ======================================================
        # CAN HEADERS A:F
        # ======================================================

        for col, header in enumerate(
            can_headers,
            start=1
        ):

            cell = ws.cell(
                row=row,
                column=col
            )

            cell.value = header
            cell.font = Font(
                bold=True,
                size=8
            )

            cell.alignment = center
            cell.border = border

        # ======================================================
        # POWER ANALYSER HEADERS H:M
        # ======================================================

        for col, header in enumerate(
            pa_headers,
            start=8
        ):

            cell = ws.cell(
                row=row,
                column=col
            )

            cell.value = header
            cell.font = Font(
                bold=True,
                size=8
            )

            cell.alignment = center
            cell.border = border

        # ======================================================
        # DATA ROWS
        # ======================================================

        row += 1

        first_data_row = row

        voltages = [
            100,
            230,
            270
        ]

        for index, voltage in enumerate(
            voltages
        ):

            data_row = first_data_row + index

            # --------------------------------------------------
            # CAN
            # --------------------------------------------------

            ws.cell(
                row=data_row,
                column=1
            ).value = voltage

            # --------------------------------------------------
            # GAP G
            # --------------------------------------------------

            ws.cell(
                row=data_row,
                column=7
            ).value = ""

            ws.cell(
                row=data_row,
                column=7
            ).border = Border()

            # --------------------------------------------------
            # POWER ANALYSER
            # --------------------------------------------------

            ws.cell(
                row=data_row,
                column=8
            ).value = voltage

            # --------------------------------------------------
            # CAN BORDER
            # --------------------------------------------------

            for col in range(1, 7):

                cell = ws.cell(
                    row=data_row,
                    column=col
                )

                cell.border = border
                cell.alignment = center

            # --------------------------------------------------
            # POWER ANALYSER BORDER
            # --------------------------------------------------

            for col in range(8, 14):

                cell = ws.cell(
                    row=data_row,
                    column=col
                )

                cell.border = border
                cell.alignment = center

        last_data_row = first_data_row + 2

        # ======================================================
        # MERGE LOAD REGULATION
        # ======================================================

        # CAN F
        ws.merge_cells(
            start_row=first_data_row,
            start_column=6,
            end_row=last_data_row,
            end_column=6
        )

        # Power Analyser M
        ws.merge_cells(
            start_row=first_data_row,
            start_column=13,
            end_row=last_data_row,
            end_column=13
        )

        # ======================================================
        # RETURN
        # ======================================================

        return last_data_row + 1

    # ==========================================================
    # CREATE HP DCDC SECTION
    # ==========================================================

    def _create_hpdcdc_section(
        self,
        row,
        hv_voltage
    ):

        ws = self.worksheet

        # ======================================================
        # STYLES
        # ======================================================

        thin = Side(
            style="thin",
            color="000000"
        )

        border = Border(
            left=thin,
            right=thin,
            top=thin,
            bottom=thin
        )

        green_fill = PatternFill(
            fill_type="solid",
            fgColor="D9EAD3"
        )

        center = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True
        )

        # ======================================================
        # SECTION INFORMATION
        # ======================================================

        section = {
            "hv_voltage": hv_voltage,
            "title_row": row,
            "condition_row": row + 1,
            "header_row": row + 2,
            "data_start": row + 3,
            "data_end": row + 5
        }

        self.hpdcdc_sections.append(section)

        # ======================================================
        # TITLE
        # ======================================================

        # CAN A:F
        ws.merge_cells(
            start_row=row,
            start_column=1,
            end_row=row,
            end_column=6
        )

        # Power Analyser H:M
        ws.merge_cells(
            start_row=row,
            start_column=8,
            end_row=row,
            end_column=13
        )

        can_title = ws.cell(
            row=row,
            column=1
        )

        can_title.value = (
            "HP DCDC Load Regulation - CAN Data"
        )

        can_title.font = Font(
            bold=True,
            size=10
        )

        can_title.fill = green_fill
        can_title.alignment = center

        pa_title = ws.cell(
            row=row,
            column=8
        )

        pa_title.value = (
            "HP DCDC Load Regulation - "
            "Power Analyser Data"
        )

        pa_title.font = Font(
            bold=True,
            size=10
        )

        pa_title.fill = green_fill
        pa_title.alignment = center

        # Borders
        for col in range(1, 7):

            cell = ws.cell(
                row=row,
                column=col
            )

            cell.border = border
            cell.fill = green_fill

        for col in range(8, 14):

            cell = ws.cell(
                row=row,
                column=col
            )

            cell.border = border
            cell.fill = green_fill

        # ======================================================
        # TEST CONDITION
        # ======================================================

        row += 1

        ws.merge_cells(
            start_row=row,
            start_column=1,
            end_row=row,
            end_column=6
        )

        ws.merge_cells(
            start_row=row,
            start_column=8,
            end_row=row,
            end_column=13
        )

        condition = (
            f"HP DCDC Set Voltage: {hv_voltage}"
        )

        can_condition = ws.cell(
            row=row,
            column=1
        )

        can_condition.value = condition
        can_condition.font = Font(
            bold=True,
            size=9
        )
        can_condition.alignment = center

        pa_condition = ws.cell(
            row=row,
            column=8
        )

        pa_condition.value = condition
        pa_condition.font = Font(
            bold=True,
            size=9
        )
        pa_condition.alignment = center

        # ======================================================
        # HEADERS
        # ======================================================

        row += 1

        can_headers = [
            "Set HV Voltage",
            "Input HV Voltage CAN(V)",
            "Input HV Current CAN(A)",
            "Output HP DCDC Voltage CAN(V)",
            "Output HP DCDC Current CAN(A)",
            "Load Regulation(%)"
        ]

        pa_headers = [
            "Set HV Voltage",
            "Input HV Voltage Power Analyser(V)",
            "Input HV Current Power Analyser(A)",
            "Output HP DCDC Voltage Power Analyser(V)",
            "Output HP DCDC Current Power Analyser(A)",
            "Load Regulation(%)"
        ]

        # ======================================================
        # CAN HEADERS
        # ======================================================

        for col, header in enumerate(
            can_headers,
            start=1
        ):

            cell = ws.cell(
                row=row,
                column=col
            )

            cell.value = header
            cell.font = Font(
                bold=True,
                size=8
            )

            cell.alignment = center
            cell.border = border

        # ======================================================
        # POWER ANALYSER HEADERS
        # ======================================================

        for col, header in enumerate(
            pa_headers,
            start=8
        ):

            cell = ws.cell(
                row=row,
                column=col
            )

            cell.value = header
            cell.font = Font(
                bold=True,
                size=8
            )

            cell.alignment = center
            cell.border = border

        # ======================================================
        # DATA ROWS
        # ======================================================

        row += 1

        first_data_row = row

        voltage_points = [
            80,
            100,
            120
        ]

        for index, voltage in enumerate(
            voltage_points
        ):

            data_row = first_data_row + index

            # CAN
            ws.cell(
                row=data_row,
                column=1
            ).value = voltage

            # GAP
            ws.cell(
                row=data_row,
                column=7
            ).value = ""

            ws.cell(
                row=data_row,
                column=7
            ).border = Border()

            # POWER ANALYSER
            ws.cell(
                row=data_row,
                column=8
            ).value = voltage

            # CAN BORDER
            for col in range(1, 7):

                cell = ws.cell(
                    row=data_row,
                    column=col
                )

                cell.border = border
                cell.alignment = center

            # POWER ANALYSER BORDER
            for col in range(8, 14):

                cell = ws.cell(
                    row=data_row,
                    column=col
                )

                cell.border = border
                cell.alignment = center

        last_data_row = first_data_row + 2

        # ======================================================
        # MERGE LOAD REGULATION
        # ======================================================

        # CAN F
        ws.merge_cells(
            start_row=first_data_row,
            start_column=6,
            end_row=last_data_row,
            end_column=6
        )

        # Power Analyser M
        ws.merge_cells(
            start_row=first_data_row,
            start_column=13,
            end_row=last_data_row,
            end_column=13
        )

        return last_data_row + 1

    # ==========================================================
    # WRITE OBC MEASUREMENT
    # ==========================================================

    def write_obc_measurement(
        self,
        row,
        set_voltage,
        can_input_voltage="",
        can_input_current="",
        can_output_voltage="",
        can_output_current="",
        pa_input_voltage="",
        pa_input_current="",
        pa_output_voltage="",
        pa_output_current=""
    ):

        ws = self.worksheet

        # ======================================================
        # CAN
        # ======================================================

        ws.cell(
            row=row,
            column=1
        ).value = set_voltage

        ws.cell(
            row=row,
            column=2
        ).value = can_input_voltage

        ws.cell(
            row=row,
            column=3
        ).value = can_input_current

        ws.cell(
            row=row,
            column=4
        ).value = can_output_voltage

        ws.cell(
            row=row,
            column=5
        ).value = can_output_current

        # ======================================================
        # GAP
        # ======================================================

        ws.cell(
            row=row,
            column=7
        ).value = ""

        ws.cell(
            row=row,
            column=7
        ).border = Border()

        # ======================================================
        # POWER ANALYSER
        # ======================================================

        ws.cell(
            row=row,
            column=8
        ).value = set_voltage

        ws.cell(
            row=row,
            column=9
        ).value = pa_input_voltage

        ws.cell(
            row=row,
            column=10
        ).value = pa_input_current

        ws.cell(
            row=row,
            column=11
        ).value = pa_output_voltage

        ws.cell(
            row=row,
            column=12
        ).value = pa_output_current

        # ======================================================
        # ALIGNMENT
        # ======================================================

        for col in list(range(1, 7)) + list(range(8, 14)):

            ws.cell(
                row=row,
                column=col
            ).alignment = Alignment(
                horizontal="center",
                vertical="center",
                wrap_text=True
            )

    # ==========================================================
    # WRITE HP DCDC MEASUREMENT
    # ==========================================================

    def write_hpdcdc_measurement(
        self,
        row,
        set_voltage,
        can_input_voltage="",
        can_input_current="",
        can_output_voltage="",
        can_output_current="",
        pa_input_voltage="",
        pa_input_current="",
        pa_output_voltage="",
        pa_output_current=""
    ):

        ws = self.worksheet

        # ======================================================
        # CAN
        # ======================================================

        ws.cell(
            row=row,
            column=1
        ).value = set_voltage

        ws.cell(
            row=row,
            column=2
        ).value = can_input_voltage

        ws.cell(
            row=row,
            column=3
        ).value = can_input_current

        ws.cell(
            row=row,
            column=4
        ).value = can_output_voltage

        ws.cell(
            row=row,
            column=5
        ).value = can_output_current

        # ======================================================
        # GAP
        # ======================================================

        ws.cell(
            row=row,
            column=7
        ).value = ""

        ws.cell(
            row=row,
            column=7
        ).border = Border()

        # ======================================================
        # POWER ANALYSER
        # ======================================================

        ws.cell(
            row=row,
            column=8
        ).value = set_voltage

        ws.cell(
            row=row,
            column=9
        ).value = pa_input_voltage

        ws.cell(
            row=row,
            column=10
        ).value = pa_input_current

        ws.cell(
            row=row,
            column=11
        ).value = pa_output_voltage

        ws.cell(
            row=row,
            column=12
        ).value = pa_output_current

        # ======================================================
        # ALIGNMENT
        # ======================================================

        for col in list(range(1, 7)) + list(range(8, 14)):

            ws.cell(
                row=row,
                column=col
            ).alignment = Alignment(
                horizontal="center",
                vertical="center",
                wrap_text=True
            )

    # ==========================================================
    # CALCULATE OBC LOAD REGULATION
    # ==========================================================

    def calculate_obc_load_regulation(
        self,
        row_100,
        row_230,
        row_270
    ):

        ws = self.worksheet

        # ======================================================
        # CAN
        # Output OBC Voltage = D
        # ======================================================

        can_formula = (
            f"=(D{row_100}-D{row_270})/"
            f"D{row_230}*100"
        )

        ws.cell(
            row=row_100,
            column=6
        ).value = can_formula

        ws.cell(
            row=row_100,
            column=6
        ).number_format = "0.00"

        # ======================================================
        # POWER ANALYSER
        # Output OBC Voltage = K
        # ======================================================

        pa_formula = (
            f"=(K{row_100}-K{row_270})/"
            f"K{row_230}*100"
        )

        ws.cell(
            row=row_100,
            column=13
        ).value = pa_formula

        ws.cell(
            row=row_100,
            column=13
        ).number_format = "0.00"

    # ==========================================================
    # CALCULATE HP DCDC LOAD REGULATION
    # ==========================================================

    def calculate_hpdcdc_load_regulation(
        self,
        row_80,
        row_100,
        row_120
    ):

        ws = self.worksheet

        # ======================================================
        # CAN
        # Output HP DCDC Voltage = D
        # ======================================================

        can_formula = (
            f"=(D{row_80}-D{row_120})/"
            f"D{row_100}*100"
        )

        ws.cell(
            row=row_80,
            column=6
        ).value = can_formula

        ws.cell(
            row=row_80,
            column=6
        ).number_format = "0.00"

        # ======================================================
        # POWER ANALYSER
        # Output HP DCDC Voltage = K
        # ======================================================

        pa_formula = (
            f"=(K{row_80}-K{row_120})/"
            f"K{row_100}*100"
        )

        ws.cell(
            row=row_80,
            column=13
        ).value = pa_formula

        ws.cell(
            row=row_80,
            column=13
        ).number_format = "0.00"

    # ==========================================================
    # GET OBC SECTION ROWS
    # ==========================================================

    def get_obc_section_rows(
        self,
        index
    ):

        if index < 0:
            raise IndexError(
                "OBC section index cannot be negative."
            )

        if index >= len(self.obc_sections):
            raise IndexError(
                "OBC section does not exist."
            )

        section = self.obc_sections[index]

        return (
            section["data_start"],
            section["data_start"] + 1,
            section["data_start"] + 2
        )

    # ==========================================================
    # GET HP DCDC SECTION ROWS
    # ==========================================================

    def get_hpdcdc_section_rows(
        self,
        index
    ):

        if index < 0:
            raise IndexError(
                "HP DCDC section index cannot be negative."
            )

        if index >= len(self.hpdcdc_sections):
            raise IndexError(
                "HP DCDC section does not exist."
            )

        section = self.hpdcdc_sections[index]

        return (
            section["data_start"],
            section["data_start"] + 1,
            section["data_start"] + 2
        )

    # ==========================================================
    # FORMAT EXCEL
    # ==========================================================

    def _format_excel(self):

        ws = self.worksheet

        # ======================================================
        # STYLES
        # ======================================================

        thin = Side(
            style="thin",
            color="000000"
        )

        thin_border = Border(
            left=thin,
            right=thin,
            top=thin,
            bottom=thin
        )

        center = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True
        )

        # ======================================================
        # ALL CELLS
        # ======================================================

        for row in ws.iter_rows():

            for cell in row:

                # G is the empty space
                if cell.column == 7:
                    continue

                if cell.value is not None:

                    cell.alignment = center
                    cell.border = thin_border

        # ======================================================
        # MAIN TITLE
        # ======================================================

        ws["A1"].font = Font(
            bold=True,
            size=16
        )

        ws["A1"].alignment = Alignment(
            horizontal="center",
            vertical="center"
        )

        # ======================================================
        # INFORMATION HEADER
        # ======================================================

        for col in range(1, 8):

            ws.cell(
                row=3,
                column=col
            ).font = Font(
                bold=True,
                size=9
            )

            ws.cell(
                row=3,
                column=col
            ).border = thin_border

            ws.cell(
                row=4,
                column=col
            ).border = thin_border

        # ======================================================
        # SECTION TITLES
        # ======================================================

        section_names = [
            "OBC Load Regulation - CAN Data",
            "OBC Load Regulation - Power Analyser Data",
            "HP DCDC Load Regulation - CAN Data",
            "HP DCDC Load Regulation - Power Analyser Data"
        ]

        for row in range(
            1,
            ws.max_row + 1
        ):

            for col in [1, 8]:

                value = ws.cell(
                    row=row,
                    column=col
                ).value

                if value in section_names:

                    ws.cell(
                        row=row,
                        column=col
                    ).font = Font(
                        bold=True,
                        size=10
                    )

        # ======================================================
        # TABLE HEADERS
        # ======================================================

        header_keywords = [
            "Input",
            "Output",
            "Load Regulation",
            "Set HV Voltage",
            "Set HV Current",
            "HV Load"
        ]

        for row in range(
            1,
            ws.max_row + 1
        ):

            for col in (
                list(range(1, 7))
                +
                list(range(8, 14))
            ):

                cell = ws.cell(
                    row=row,
                    column=col
                )

                if cell.value is None:
                    continue

                value = str(
                    cell.value
                )

                if any(
                    keyword in value
                    for keyword in header_keywords
                ):

                    cell.font = Font(
                        bold=True,
                        size=8
                    )

        # ======================================================
        # COLUMN WIDTHS
        # ======================================================

        widths = {

            # CAN
            "A": 14,
            "B": 17,
            "C": 20,
            "D": 22,
            "E": 22,
            "F": 20,

            # GAP
            "G": 8,

            # POWER ANALYSER
            "H": 14,
            "I": 22,
            "J": 22,
            "K": 24,
            "L": 24,
            "M": 20
        }

        for column, width in widths.items():

            ws.column_dimensions[
                column
            ].width = width

        # ======================================================
        # ROW HEIGHT
        # ======================================================

        for row in range(
            1,
            ws.max_row + 1
        ):

            ws.row_dimensions[
                row
            ].height = 28

        ws.row_dimensions[1].height = 32

        ws.row_dimensions[3].height = 32
        ws.row_dimensions[4].height = 30

        # ======================================================
        # OBC HEADER HEIGHT
        # ======================================================

        for section in self.obc_sections:

            ws.row_dimensions[
                section["header_row"]
            ].height = 60

            for r in range(
                section["data_start"],
                section["data_end"] + 1
            ):

                ws.row_dimensions[r].height = 30

        # ======================================================
        # HP DCDC HEADER HEIGHT
        # ======================================================

        for section in self.hpdcdc_sections:

            ws.row_dimensions[
                section["header_row"]
            ].height = 65

            for r in range(
                section["data_start"],
                section["data_end"] + 1
            ):

                ws.row_dimensions[r].height = 30

        # ======================================================
        # FREEZE PANES
        # ======================================================

        ws.freeze_panes = "A6"

        # ======================================================
        # GRID
        # ======================================================

        ws.sheet_view.showGridLines = True

        # ======================================================
        # PRINT SETTINGS
        # ======================================================

        ws.page_setup.orientation = "landscape"

        ws.page_setup.fitToWidth = 1

        ws.page_setup.fitToHeight = 0

        ws.sheet_properties.pageSetUpPr.fitToPage = True

    # ==========================================================
    # SAVE REPORT
    # ==========================================================

    def save(self):

        if self.workbook is None:

            raise RuntimeError(
                "Report has not been started."
            )

        self._format_excel()

        self.workbook.save(
            self.report_path
        )

        print(
            "\nReport saved:"
        )

        print(
            self.report_path
        )

        return self.report_path

